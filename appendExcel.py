#coding = utf-8

from openpyxl import Workbook, load_workbook

def appendNewExcel(file_appended_to, file_to_append):
    
    #读取需合并的Excel文件，获取默认活动工作表，并获得该工作表的行数和列数
    wb = load_workbook(filename = file_to_append)
    ws = wb.get_active_sheet()
    rows = ws.get_highest_row()
    columns = ws.get_highest_column()

    #读取合并至的Excel文件，获取默认活动工作表
    wb2 = load_workbook(filename = file_appended_to)
    ws2 = wb2.get_active_sheet()

    #循环读取需合并的Excel工作表的单元格并以行为单位建立单元格的列表
    for row in range(rows):
        rowlist = []
        for column in range(columns):
            rowlist.append(ws.cell(row = row, column = column).value)

        #将单元格的列表添加到目标Excel文件的工作表中
        ws2.append(rowlist)

    #保存对目标Excel文件的更改
    wb2.save(file_appended_to)


#仅仅为了测试使用的代码，打印Excel默认活动工作表的所有单元格
#def printExcel():
#    wb = load_workbook(filename = r'D:\Documents\论文数据.xlsx')
#    ws = wb.get_active_sheet()
#    rows = ws.get_highest_row()
#    columns = ws.get_highest_column()
#
#    for row in range(rows):
#        for column in range(columns):
#            print(ws.cell(row = row, column = column).value, end = ' ')
#            print('\n')
